from pathlib import Path

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models.course_registration import CourseRegistration
from app.models.contact_message import ContactMessage
from app.models.internship_application import InternshipApplication
from app.services.google_drive_service import upload_excel_to_google_drive

EXCEL_DIR = Path("app/exports")
EXCEL_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_FILE = EXCEL_DIR / "QodeKraft_Management.xlsx"


def export_all_data(db: Session):
    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # ============================================================
    # COURSE REGISTRATIONS
    # ============================================================

    course_sheet = workbook.create_sheet("Course Registrations")

    registrations = (
        db.query(CourseRegistration)
        .order_by(CourseRegistration.id.asc())
        .all()
    )

    course_headers = [
        "Registration ID",
        "Full Name",
        "Email",
        "Phone",
        "Referral ID",
        "Course ID",
        "Course Slug",
        "Course Title",
        "Amount",
        "Razorpay Order ID",
        "Razorpay Payment ID",
        "UTR",
        "Payment Status",
        "Created At",
        "Paid At",
    ]

    course_sheet.append(course_headers)

    for registration in registrations:
        course_sheet.append([
            registration.registration_id,
            registration.full_name,
            registration.email,
            registration.phone,
            registration.referral_id,
            registration.course_id,
            registration.course_slug,
            registration.course_title,
            registration.amount,
            registration.razorpay_order_id,
            registration.razorpay_payment_id,
            registration.utr,
            registration.payment_status,
            registration.created_at,
            registration.paid_at,
        ])

    # ============================================================
    # CONTACT MESSAGES
    # ============================================================

    contact_sheet = workbook.create_sheet("Contact Messages")

    contacts = (
        db.query(ContactMessage)
        .order_by(ContactMessage.id.asc())
        .all()
    )

    contact_columns = [
        column.name
        for column in ContactMessage.__table__.columns
    ]

    contact_sheet.append(contact_columns)

    for contact in contacts:
        contact_sheet.append([
            getattr(contact, column)
            for column in contact_columns
        ])

    # ============================================================
    # INTERNSHIP APPLICATIONS
    # ============================================================

    internship_sheet = workbook.create_sheet("Internship Applications")

    internships = (
        db.query(InternshipApplication)
        .order_by(InternshipApplication.id.asc())
        .all()
    )

    internship_columns = [
        column.name
        for column in InternshipApplication.__table__.columns
    ]

    internship_sheet.append(internship_columns)

    for internship in internships:
        internship_sheet.append([
            getattr(internship, column)
            for column in internship_columns
        ])

   # ============================================================
# SAVE EXCEL
# ============================================================

    workbook.save(EXCEL_FILE)

# Upload the updated Excel to Google Drive
    upload_excel_to_google_drive(EXCEL_FILE)

    return EXCEL_FILE